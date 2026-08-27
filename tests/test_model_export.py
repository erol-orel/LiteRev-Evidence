"""Tests for model_export.build_model_xlsx — the Excel export of a model bundle.

Skips cleanly where openpyxl isn't installed (it is a runtime dep in requirements.txt).
"""
import io

import pytest

pytest.importorskip("openpyxl")
from openpyxl import load_workbook

import model_export


def _bundle():
    return {
        "model_spec": {
            "outcome": {"name": "Influenza-A activity", "machine_name": "flu_a_load",
                        "task_type": "regression", "unit": "gc/day", "source": "public_api",
                        "provenance": [11, 22]},
            "features": [
                {"name": "RSV activity", "machine_name": "rsv_load", "dtype": "float",
                 "source": "public_api", "importance": "medium", "provenance": [33]},
                {"name": "Season (cos)", "machine_name": "season_cos", "dtype": "float",
                 "source": "user", "importance": "high"},
            ],
        },
        "dataset": {
            "columns": ["date", "flu_a_load", "rsv_load", "season_cos"],
            "rows": [
                {"date": "2023-01-30", "flu_a_load": 541443.9, "rsv_load": 2.0e7, "season_cos": 0.55},
                {"date": "2023-02-06", "flu_a_load": 2888804.8, "rsv_load": 2.4e7, "season_cos": 0.64},
            ],
        },
        "runs": [{"run_id": 1, "is_active": True, "family": "random_forest",
                  "task_type": "regression", "metric": "r2",
                  "metrics": {"r2": 0.51, "rmse": 8.3e6},
                  "hyperparameters": {"n_estimators": 387, "max_depth": 13},
                  "trained_at": "2026-07-31T10:00:00Z"}],
    }


def test_xlsx_has_all_sheets_with_values_and_outcome():
    wb = load_workbook(io.BytesIO(model_export.build_model_xlsx(_bundle())))
    assert wb.sheetnames == ["Variables", "Dataset", "Model runs", "SEIR parameters"]

    variables = wb["Variables"]
    assert [c.value for c in variables[1]][:3] == ["role", "name", "machine_name"]
    # outcome first, then features; provenance ids joined
    assert [variables[2][0].value, variables[2][2].value, variables[2][7].value] == ["outcome", "flu_a_load", "11, 22"]
    assert variables[3][0].value == "feature" and variables[3][2].value == "rsv_load"

    dataset = wb["Dataset"]
    assert [c.value for c in dataset[1]] == ["date", "flu_a_load", "rsv_load", "season_cos"]  # every variable + outcome
    assert dataset[2][1].value == 541443.9                     # a real value round-trips
    assert dataset.max_row == 3                                 # header + 2 observations

    runs = wb["Model runs"]
    assert runs[2][2].value == "random_forest"
    assert "n_estimators" in runs[2][6].value                  # hyperparameters serialised


def test_xlsx_handles_empty_bundle_without_crashing():
    wb = load_workbook(io.BytesIO(model_export.build_model_xlsx({})))
    assert wb.sheetnames == ["Variables", "Dataset", "Model runs", "SEIR parameters"]
    assert wb["Dataset"].cell(row=1, column=1).value.startswith("(no dataset")


def test_xlsx_includes_the_seir_block_with_provenance():
    """The epidemiological parameters used to be dropped from the export entirely."""
    b = _bundle()
    b["model_spec"]["epidemic_parameters"] = {
        "applicable": True, "disease": "Influenza (seasonal)",
        "params": {
            "r0": {"value": 1.3, "ci_low": 1.2, "ci_high": 1.4, "unit": "",
                   "n_studies": 4, "provenance": [11, 22]},
            "cfr": {"value": 0.001, "unit": "proportion", "n_studies": 2,
                    "provenance": [33], "overridden": True},
        },
    }
    ws = load_workbook(io.BytesIO(model_export.build_model_xlsx(b)))["SEIR parameters"]
    rows = {r[0].value: r for r in ws.iter_rows(min_row=2) if r[0].value}
    assert [c.value for c in ws[1]][:4] == ["parameter", "value", "ci_low", "ci_high"]
    assert rows["r0"][1].value == 1.3 and rows["r0"][6].value == "11, 22"
    assert rows["cfr"][7].value == "yes"          # user override is flagged
    assert rows["disease"][1].value == "Influenza (seasonal)"


def test_xlsx_seir_sheet_says_so_when_there_is_no_seir_model():
    ws = load_workbook(io.BytesIO(model_export.build_model_xlsx(_bundle())))["SEIR parameters"]
    assert "no SEIR model" in str(ws.cell(row=2, column=1).value)
