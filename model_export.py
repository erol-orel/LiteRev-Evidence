"""Build a spreadsheet (.xlsx) export of a scenario's model from an export bundle.

Pure and importable without the web app: `main.export_model_xlsx` assembles the same
bundle as the JSON export (`export_model_bundle`) and hands it here to serialise. Kept
separate so the workbook layout is unit-testable without FastAPI. openpyxl is imported
lazily so importing this module never fails when the optional dependency is absent.

Sheets:
  - Variables : every model variable (the outcome + each feature) with its role, name,
    machine name, dtype/task, unit, source and provenance article ids.
  - Dataset   : the actual data — one column per variable + the outcome, one row per
    observation (the values used to train / score the model).
  - Model runs: every trained run with its metrics and hyperparameters.
"""
from __future__ import annotations

import json
from typing import Any


def _prov(d: dict) -> str:
    ids = d.get("provenance") or []
    return ", ".join(str(x) for x in ids if x is not None)


def _cell(v: Any):
    """Coerce a value to something openpyxl can write (str/int/float/None)."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return json.dumps(v, ensure_ascii=False, default=str)


def build_model_xlsx(bundle: dict) -> bytes:
    """Serialise an export bundle (see main.export_model_bundle) to .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io

    bundle = bundle or {}
    spec = bundle.get("model_spec") or {}
    outcome = spec.get("outcome") or {}
    features = spec.get("features") or []
    dataset = bundle.get("dataset") or {}
    runs = bundle.get("runs") or []
    bold = Font(bold=True)

    wb = Workbook()

    # ── Sheet 1: Variables ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Variables"
    header = ["role", "name", "machine_name", "dtype / task", "unit", "source", "importance",
              "provenance_article_ids"]
    ws.append(header)
    for c in ws[1]:
        c.font = bold
    if outcome:
        ws.append(["outcome", _cell(outcome.get("name")), _cell(outcome.get("machine_name")),
                   _cell(outcome.get("task_type")), _cell(outcome.get("unit")),
                   _cell(outcome.get("source")), "", _prov(outcome)])
    for f in features:
        ws.append(["feature", _cell(f.get("name")), _cell(f.get("machine_name")),
                   _cell(f.get("dtype")), _cell(f.get("unit")), _cell(f.get("source")),
                   _cell(f.get("importance")), _prov(f)])

    # ── Sheet 2: Dataset (all variable values + the outcome) ───────────────────
    ws2 = wb.create_sheet("Dataset")
    rows = dataset.get("rows") or []
    cols = dataset.get("columns") or (list(rows[0].keys()) if rows else [])
    cols = [str(c) for c in cols]
    if cols:
        ws2.append(cols)
        for c in ws2[1]:
            c.font = bold
        for r in rows:
            ws2.append([_cell(r.get(c)) for c in cols])
    else:
        ws2.append(["(no dataset attached to this scenario)"])

    # ── Sheet 3: Model runs (metrics + hyperparameters) ───────────────────────
    ws3 = wb.create_sheet("Model runs")
    run_header = ["run_id", "active", "family", "task_type", "metric", "metrics",
                  "hyperparameters", "trained_at"]
    ws3.append(run_header)
    for c in ws3[1]:
        c.font = bold
    for r in runs:
        ws3.append([
            _cell(r.get("run_id")), "yes" if r.get("is_active") else "",
            _cell(r.get("family")), _cell(r.get("task_type")), _cell(r.get("metric")),
            _cell(r.get("metrics")), _cell(r.get("hyperparameters")), _cell(r.get("trained_at")),
        ])

    # Reasonable column widths for readability.
    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
